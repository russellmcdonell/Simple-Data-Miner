#!/usr/bin/env python

# pylint: disable=invalid-name, line-too-long, pointless-string-statement, unspecified-encoding
# pylint: disable=broad-exception-caught, unused-variable, too-many-lines, abstract-class-instantiated

'''
A script to build a web site where users, with little or no SQL knowledge, can mine a database.

This script lets users select a table for data mining, select their column, constrain their column
and, if they want, count and sum their columns, into an extract that they can download as an Excel workbook.

    SYNOPSIS
    $ export FLASK_APP=SimpleDataMiner
    $ python3 -m flask run
        [-D DatabaseType|--DatabaseType=DatabaseType]
        [-I inputDir|--inputDir=inputDir]
        [-i inputWorkbook|--inputWorkbook=inputWorkbook]
        [-C configDir|--configDir=configDir]
        [-c configFile|--configFile=configFile]
        [-s server|--server=server]
        [-u username|--username=username]
        [-p password|--password=password]
        [-d databaseName|--databaseName=databaseName]
        [-v loggingLevel|--verbose=logingLevel]
        [-L logDir|--logDir=logDir]
        [-l logfile|--logfile=logfile]

    REQUIRED
    -D DatabaseType|--DatabaseType=DatabaseType
    The type of database [choice:MSSQL/MySQL]


    OPTIONS  
    -I inputDir|--inputDir=inputDir
    The directory containing the Excel workbook which contains
    the configuration of the minable tables

    -i inputWorkbook|--inputWorkbook=inputWorkbook
    The Excel workbook which contains
    the configuration of the minable tables

    -C configDir|--configDir=configDir
    The directory containing the database connection configuration file
    (default='databaseConfig')

    -c configFile|--configFile=configFile
    The database connection configuration file (default=SimpleDataMiner.json)
    which has the default database values for each Database Type.
    These can be overwritten using command line options.

    -s server|--server=server]
    The address of the database server

    -u userName|--userName=userName]
    The user name require to access the database

    -p password|--userName=userName]
    The user password require to access the database

    -d databaseName|--databaseName=databaseName]
    The name of the database

    -v loggingLevel|--verbose=loggingLevel
    Set the level of logging that you want.

    -O logDir|--logDir=logDir
    The directory where the log file will be created
    (default=".").

    -o logfile|--logfile=logfile
    The name of a log file where you want all messages captured.


    THE MAIN CODE
    Start by parsing the command line arguements, setting up logging and checking connectivity to the database.
    Then check the Excel workbook - it should have one sheet of database table names/worksheet names pairs.
    Then one sheet per database table of database table configuration.

'''

# Import all the modules that make life easy
import io
import sys
import os
import argparse
import logging
import collections
import json
import ast
import dateutil.parser
import dateutil.tz
import pandas as pd
from sqlalchemy import MetaData, create_engine, text
from sqlalchemy.orm import sessionmaker
from sqlalchemy.exc import OperationalError
from sqlalchemy_utils import database_exists
from flask import Flask, url_for, request, send_file, Response
from openpyxl import load_workbook
import data as d


app = Flask(__name__)

def convertInWeb(thisValue):
    '''
    Convert a value from a web form
    '''
    if not isinstance(thisValue, str):
        return thisValue
    if thisValue == '':
        return None
    try:
        newValue = ast.literal_eval(thisValue)
    except Exception as thisE:
        newValue = thisValue
    return newValue


@app.route('/', methods=['GET'])
def splash():
    '''
    Display the Welcome splash page
    '''

    message = '<html><head><title>Simple Data Miner</title><link rel="icon" href="data:,"></head><body style="font-size:120%">'
    message += '<h1 style="text-align:center">Welcolme to the Simple Data Miner</h1>'
    message += '<h2 style="text-align:center">Please select the data table you wish to mine</h2>'
    message += f'<form id="tables" action ="{url_for("doSelectColumns")}" method="post" enctype="multipart/form-data" style="font-size:120%">'
    message += '<select name="table" style="font-size:120%">'
    for mineTable, tableConfig in  d.mineTables.items():
        message += f'<option value="{mineTable}">{tableConfig["tableName"]}'
    message += '</select><br/><br/>'
    message += '<input id="submit" type="submit" name="submit" value="Please mine this table" style="font-size:120%">'
    message += '</form>'
    message += '</body></html>'
    return Response(response=message, status=200)


@app.route('/doSelectColumns', methods=['POST'])
def doSelectColumns():
    '''
    For the selected table, list the columns and ask the user to select which ones are to be included in the extract
    '''
    message = '<html><head><title>Simple Data Miner</title><link rel="icon" href="data:,"></head><body style="font-size:120%">'
    message += '<h1 style="text-align:center">Simple Data Miner</h1>'
    thisMessage, thisTable, dummy1, dummy2, dummy3, dummy4 = checkForm(request, 1)
    if thisMessage is not None:
        message += thisMessage
        return Response(response=message, status=400)
    message += f'<h2 style="text-align:center">For the "{d.mineTables[thisTable]["tableName"]}" table</h2>'
    message += '<h3 style="text-align:center">Please select the columns you would like mined into your extract</h3>'
    message += f'<form id="selected" action ="{url_for("constrainColumns")}" method="post" enctype="multipart/form-data">'
    message += f'<input id="table" type="hidden" name="table" value="{thisTable}">'
    message += '<table>'
    for i, thisCol in enumerate(d.mineTables[thisTable]['columns']):
        message += f'<tr><td><input type="checkbox" name="selected" value="{i}"></td><td style="font-size:150%">{thisCol["columnName"]}</td></tr>'
    message += '</table>'
    message += '<br/>'
    message += f'<input id="submit" type="submit" name="submit" value="Please mine these columns in the {d.mineTables[thisTable]["tableName"]} table" style="font-size:120%">'
    message += '</form>'
    message += '</body></html>'
    return Response(response=message, status=200)


@app.route('/constrainColumns', methods=['POST'])
def constrainColumns():
    '''
    List the selected columns and let the user select any columns that they would like to constrain
    '''
    message = '<html><head><title>Simple Data Miner</title><link rel="icon" href="data:,"></head><body style="font-size:120%">'
    message += '<h1 style="text-align:center">Simple Data Miner</h1>'
    thisMessage, thisTable, dummy1, dummy2, dummy3, dummy4 = checkForm(request, 1)
    if thisMessage is not None:
        message += thisMessage
        return Response(response=message, status=400)
    if 'selected' not in request.form:
        message += f'<form id="tables" action ="{url_for("selectColumns")}" method="post" enctype="multipart/form-data">'
        message += f'<input id="table" type="hidden" name="table" value="{thisTable}">'
        message += '<input id="submit" type="submit" name="select" value="No columns selected - click here to try again" style="font-size:120%">'
        message += '</form>'
        message += '</body></html>'
        return Response(response=message, status=400)
    message += f'<h2 style="text-align:center">For the "{d.mineTables[thisTable]["tableName"]}" table</h2>'
    message += '<h3 style="text-align:center">Please select any columns that you would like constrained in you mined extract</h3>'
    message += f'<form id="selected" action ="{url_for("doNextConstraint")}" method="post" enctype="multipart/form-data">'
    message += f'<input id="table" type="hidden" name="table" value="{thisTable}">'
    message += '<input id="first" type="hidden" name="first" value="1">'
    message += '<table>'
    columnsSelected = []
    for i, selected in enumerate(request.form.getlist('selected')):
        columnsSelected.append(convertInWeb(selected.strip()))
        message += f'<tr><td><input type="checkbox" name="selected" value="{i}"></td>'
        message += f'<td style="font-size:150%">{d.mineTables[thisTable]["columns"][int(selected)]["columnName"]}</td></tr>'
    message += '</table>'
    message += '<br/>'
    message += f'<input id="columnsSelected" type="hidden" name="columnsSelected" value="{columnsSelected}">'
    message += f'<input id="submit" type="submit" name="submit" value="Please constrain these columns when mining the {d.mineTables[thisTable]["tableName"]} table" style="font-size:120%">'
    message += '</form>'
    message += '</body></html>'
    return Response(response=message, status=200)


def makeConstraint(thisTable, columnsSelected, constrainedColumns, nextConstraint, where):
    '''
    Build the web form for selecting a constraint
    '''
    thisColumn = columnsSelected[int(constrainedColumns[nextConstraint])]
    thisCol = d.mineTables[thisTable]['columns'][thisColumn]
    thisColumnName = thisCol['columnName']
    thisDatatype = thisCol['datatype']
    thisColumnLookup = thisCol['lookupTable']
    thisColumnLookupCode = thisCol['lookupCodeColumn']
    thisColumnLookupDesc = thisCol['lookupDescriptionColumn']
    message = f'<h2 style="text-align:center">For the column "{thisColumnName}" in the "{d.mineTables[thisTable]["tableName"]}" table</h2>'
    if (thisDatatype != "string") or (thisColumnLookup is None):
        message += f'<h3 style="text-align:center">Please select the type of constraint(s) on the data from the "{thisColumnName}" column to restrict the data in your mined extract</h3>'
    else:
        message += f'<h3 style="text-align:center">Please select codes from the "{thisColumnName}" column that you would like included in your mined extract</h3>'
    message += f'<form id="selected" action ="{url_for("doThisConstraint")}" method="post" enctype="multipart/form-data">'
    message += f'<input id="table" type="hidden" name="table" value="{thisTable}">'
    message += f'<input id="columnsSelected" type="hidden" name="columnsSelected" value="{columnsSelected}">'
    message += f'<input id="constrainedColumn" type="hidden" name="constrainedColumns" value="{constrainedColumns}">'
    message += f'<input id="nextConstraint" type="hidden" name="nextConstraint" value="{nextConstraint}">'
    message += f"<input id='where' type='hidden' name='where' value='{where}'>"
    message += '<table>'
    if (thisDatatype != "string") or (thisColumnLookup is None):
        message += '<tr><td><input id="equals" type="checkbox" name="constraint" value="equals"></td><td style="font-size:150%">Equals a specific value</td></tr>'
        message += '<tr><td><input id="notEquals" type="checkbox" name="constraint" value="notEquals"></td><td style="font-size:150%">Does not equal a specific value</td></tr>'
        if thisDatatype != 'string':
            message += '<tr><td><input id="gtThan" type="checkbox" name="constraint" value="gtThan"></td><td style="font-size:150%">Greater than a specific value</td></tr>'
            message += '<tr><td><input id="gteThan" type="checkbox" name="constraint" value="gteThan"></td><td style="font-size:150%">Greater than or equal to a specific value</td></tr>'
            message += '<tr><td><input id="ltThan" type="checkbox" name="constraint" value="ltThan"></td><td style="font-size:150%">Less than a specif value</td></tr>'
            message += '<tr><td><input id="lteThan" type="checkbox" name="constraint" value="lteThan"></td><td style="font-size:150%">Less than or equal to a specif value</td></tr>'
            message += '<tr><td><input id="inRange" type="checkbox" name="constraint" value="inRange"></td><td style="font-size:150%">Within a range of values</td></tr>'
        else:
            message += '<tr><td><input id="starts" type="checkbox" name="constraint" value="starts"></td><td style="font-size:150%">Starts with specific string of characters</td></tr>'
            message += '<tr><td><input id="ends" type="checkbox" name="constraint" value="ends"></td><td style="font-size:150%">Ends with specific string of characters</td></tr>'
            message += '<tr><td><input id="contains" type="checkbox" name="constraint" value="contains"></td><td style="font-size:150%">Contains a specific string of characters</td></tr>'
            message += '<tr><td><input id="notContains" type="checkbox" name="constraint" value="notContains"></td><td style="font-size:150%">Does not contains a specific string of characters</td></tr>'
    else:
        selectText = f'SELECT {thisColumnLookupCode}, {thisColumnLookupDesc} FROM {thisColumnLookup}'
        codes_df = pd.read_sql_query(text(selectText), d.engine.connect())
        codes = codes_df.values.tolist()            # Convert to a list of lists (will be [[code, description]])
        for codeRow in codes:
            message += f'<tr><td><input type="checkbox" name="selectCode" value="{codeRow[0]}"></td><td style="font-size:150%">{codeRow[0]}</td><td style="font-size:150%">{codeRow[1]}</td></tr>'
    message += '</table>'
    message += '<br/>'
    if (thisDatatype != "string") or (thisColumnLookup is None):
        message += f'<input id="submit" type="submit" name="submit" value="Please apply this/these constrains to the \'{thisColumnName}\' column when mining the \'{d.mineTables[thisTable]["tableName"]}\' table" style="font-size:150%">'
    else:
        message += f'<input id="submit" type="submit" name="submit" value="Please only include these values from the \'{thisColumnName}\' column when mining the \'{d.mineTables[thisTable]["tableName"]}\' table" style="font-size:150%">'
    message += '</form>'
    message += '</body></html>'
    return message


def checkForm(thisRequest, level):
    '''
    Check that the form data hasn't go lost
    '''
    if ('table' not in thisRequest.form) or ((thisTable := convertInWeb(thisRequest.form['table'].strip())) not in d.mineTables):
        message = f'<p style="text-align:centre"><b><a href="{url_for("splash")}">Internal error (lost selected table) - please click here to start again</a></b>'
        message += '</body></html>'
        return message, thisTable, None, None, None, None
    if level == 1:
        if 'first' in request.form:
            return None, thisTable, True, None, None, None
        else:
            return None, thisTable, False, None, None, None
    if 'where' not in thisRequest.form:
        message = f'<p style="text-align:centre"><b><a href="{url_for("splash")}">Internal error (lost list of constrained columns) - please try again</a></b>'
        message += '</body></html>'
        return message, None, None, None, None, None
    where = convertInWeb(thisRequest.form['where'].strip())
    if level in [2, 3]:
        if 'columnsSelected' not in thisRequest.form:
            message = f'<p style="text-align:centre"><b><a href="{url_for("splash")}">Internal error (lost selected columns) - please try again</a></b>'
            message += '</body></html>'
            return message, None, None, None, None, None
        columnsSelected = convertInWeb(thisRequest.form['columnsSelected'].strip())
        if level == 3:
            return None, thisTable, columnsSelected, None, None, where
        if ('constrainedColumns' not in thisRequest.form) or ('nextConstraint' not in thisRequest.form):
            message = f'<p style="text-align:centre"><b><a href="{url_for("splash")}">Internal error (lost list of constrained columns) - please try again</a></b>'
            message += '</body></html>'
            return message, None, None, None, None, None
        constrainedColumns = convertInWeb(thisRequest.form['constrainedColumns'].strip())
        nextConstraint = convertInWeb(thisRequest.form['nextConstraint'].strip())
        return None, thisTable, columnsSelected, constrainedColumns, nextConstraint, where
    # Level 4
    if 'selectColumns' not in thisRequest.form:
        message = f'<p style="text-align:centre"><b><a href="{url_for("splash")}">Internal error (lost column selection) - please click here to start again</a></b>'
        message += '</body></html>'
        return message, None, None, None, None, None
    selectColumns = convertInWeb(thisRequest.form['selectColumns'].strip())
    if 'groupByColumns' not in thisRequest.form:
        message = f'<p style="text-align:centre"><b><a href="{url_for("splash")}">Internal error (lost counting/summing) - please click here to start again</a></b>'
        message += '</body></html>'
        return message, None, None, None, None, None
    groupByColumns = convertInWeb(thisRequest.form['groupByColumns'].strip())
    return None, thisTable, selectColumns, groupByColumns, None, where


@app.route('/doNextConstraint', methods=['POST'])
def doNextConstraint():
    '''
    List the next selected columns column to constrain and get the constraint type
    '''
    message = '<html><head><title>Simple Data Miner</title><link rel="icon" href="data:,"></head><body style="font-size:120%">'
    message += '<h1 style="text-align:center">Simple Data Miner</h1>'
    thisMessage, thisTable, first, dummy2, dummy3, dummy4 = checkForm(request, 1)
    if thisMessage is not None:
        message += thisMessage
        return Response(response=message, status=400)
    if 'columnsSelected' not in request.form:
        message += f'<p style="text-align:centre"><b><a href="{url_for("splash")}">Internal error (lost selected columns) - please click here to start again</a></b>'
        message += '</body></html>'
        return Response(response=message, status=400)
    columnsSelected = convertInWeb(request.form['columnsSelected'].strip())
    if 'selected' in request.form:      # First time through and something selected
        where = ''
        constrainedColumns = []
        for selected in request.form.getlist('selected'):
            constrainedColumns.append(convertInWeb(selected.strip()))
        nextConstraint = 0
    elif ('constrainedColumns' not in request.form) or ('nextConstraint' not in request.form) or ('where' not in request.form):
        if first:   # First time through and nothing selected
            where = ''
            message += buildAggs(thisTable, columnsSelected, where)
        else:
            message += f'<p style="text-align:centre"><b><a href="{url_for("splash")}">Internal error (lost list of constrained columns) - please try again</a></b>'
            message += '</body></html>'
        return Response(response=message, status=400)
    else:
        constrainedColumns = convertInWeb(request.form['constrainedColumns'].strip())
        nextConstraint = convertInWeb(request.form['nextConstraint'].strip())
        where = convertInWeb(request.form['where'].strip())
    message += makeConstraint(thisTable, columnsSelected, constrainedColumns, nextConstraint, where)
    return Response(response=message, status=200)


@app.route('/doThisConstraint', methods=['POST'])
def doThisConstraint():
    '''
    Handle the requested constraint(s) for this columns
    '''
    message = '<html><head><title>Simple Data Miner</title><link rel="icon" href="data:,"></head><body style="font-size:120%">'
    message += '<h1 style="text-align:center">Simple Data Miner</h1>'
    thisMessage, thisTable, columnsSelected, constrainedColumns, nextConstraint, where = checkForm(request, 2)
    if thisMessage is not None:
        message += thisMessage
        return Response(response=message, status=400)
    thisCol = d.mineTables[thisTable]['columns'][columnsSelected[constrainedColumns[nextConstraint]]]
    thisColumn = thisCol['column']
    thisColumnName = thisCol['columnName']
    if ('constraint' not in request.form) and ('selectCode' not in request.form):
        message += f'<form id="tables" action ="{url_for("doNextConstraint")}" method="post" enctype="multipart/form-data">'
        message += f'<input id="table" type="hidden" name="table" value="{thisTable}">'
        message += f'<input id="columnsSelected" type="hidden" name="columnsSelected" value="{columnsSelected}">'
        message += f'<input id="constrainedColumn" type="hidden" name="table" value="{constrainedColumns}">'
        message += f'<input id="nextConstraint" type="hidden" name="table" value="{nextConstraint}">'
        message += f"<input id='where' type='hidden' name='where' value='{where}'>"
        message += '<input id="submit" type="select" name="select" value="No constraint type selected - click here to try again">'
        message += '</form>'
        message += '</body></html>'
        return Response(response=message, status=400)
    if 'constraint' in request.form:
        message = f'<h2 style="text-align:center">For column "{thisColumnName}" in table "{thisTable}"</h2>'
        message = '<h3 style="text-align:center">Enter the value(s) required for this/these constraint(s)</h3>'
        message += f'<form id="setConstraints" action ="{url_for("setConstraints")}" method="post" enctype="multipart/form-data">'
        message += f'<input id="table" type="hidden" name="table" value="{thisTable}">'
        message += f'<input id="columnsSelected" type="hidden" name="columnsSelected" value="{columnsSelected}">'
        message += f"<input id='where' type='hidden' name='where' value='{where}'>"
        message += '<table>'
        constraintType = []
        for thisConstraintType in request.form.getlist('constraint'):
            constraintType.append(convertInWeb(thisConstraintType.strip()))
        message += buildConstraintValues(thisTable, columnsSelected, constrainedColumns, nextConstraint, where, constraintType)
        return Response(response=message, status=200)
    else:
        if where is None:
            where = ''
        if where != '':
            where += ' AND '
        where += f'{thisColumn} in ('
        first = True
        for thisCode in request.form.getlist('selectCode'):
            if first:
                first = False
            else:
                where += ', '
            where += f'"{thisCode}"'
        where += ')'
        nextConstraint += 1
        if nextConstraint < len(constrainedColumns):
            message += makeConstraint(thisTable, columnsSelected, constrainedColumns, nextConstraint, where)
        else:
            message += buildAggs(thisTable, columnsSelected, where)
        return Response(response=message, status=200)


def buildConstraintValues(thisTable, columnsSelected, constrainedColumns, nextConstraint, where, constraintType):
    '''
    Build the form for inputting the constraint value(s)
    '''
    thisCol = d.mineTables[thisTable]['columns'][columnsSelected[constrainedColumns[nextConstraint]]]
    thisColumnName = thisCol['columnName']
    message = f'<h2 style="text-align:center">For column "{thisColumnName}" in table "{thisTable}"</h2>'
    message += '<h3 style="text-align:center">Enter the value(s) required for this/these constraint(s)</h3>'
    message += f'<form id="setConstraints" action ="{url_for("setConstraints")}" method="post" enctype="multipart/form-data">'
    message += f'<input id="table" type="hidden" name="table" value="{thisTable}">'
    message += f'<input id="columnsSelected" type="hidden" name="columnsSelected" value="{columnsSelected}">'
    message += f'<input id="constrainedColumn" type="hidden" name="constrainedColumns" value="{constrainedColumns}">'
    message += f'<input id="nextConstraint" type="hidden" name="nextConstraint" value="{nextConstraint}">'
    message += f"<input id='where' type='hidden' name='where' value='{where}'>"
    message += '<table>'
    for thisConstraint in constraintType:
        if thisConstraint == 'equals':
            message += f'<tr><td style="font-size:150%">Enter the value that data from column "{thisColumnName}" must equal</td>'
            message += '<td><input id="input" type="text" name="inputEquals"></td></tr>'
        elif thisConstraint == 'notEquals':
            message += f'<tr><td style="font-size:150%">Enter the value that data from column "{thisColumnName}" must <b>not</b> equal</td>'
            message += '<td><input id="input" type="text" name="inputNotEquals"></td></tr>'
        elif thisConstraint == 'gtThan':
            message += f'<tr><td style="font-size:150%">Enter the value that data from column "{thisColumnName}" must be greater than</td>'
            message += '<td><input id="input" type="text" name="inputGtThan"></td></tr>'
        elif thisConstraint == 'gteThan':
            message += f'<tr><td style="font-size:150%">Enter the value that data from column "{thisColumnName}" must be equal or greater than</td>'
            message += '<td><input id="input" type="text" name="inputGteThan"></td></tr>'
        elif thisConstraint == 'ltThan':
            message += f'<tr><td style="font-size:150%">Enter the value that data from column "{thisColumnName}" must be less than</td>'
            message += '<td><input id="input" type="text" name="inputLtThan"></td></tr>'
        elif thisConstraint == 'lteThan':
            message += f'<tr><td style="font-size:150%">Enter the value that data from column "{thisColumnName}" must be equal or less than</td>'
            message += '<td><input id="input" type="text" name="inputLteThan"></td></tr>'
        elif thisConstraint == 'starts':
            message += f'<tr><td style="font-size:150%">Enter the characters that data from column "{thisColumnName}" must start with</td>'
            message += '<td><input id="input" type="text" name="inputStarts"></td></tr>'
        elif thisConstraint == 'ends':
            message += f'<tr><td style="font-size:150%">Enter the characters that data from "{thisColumnName}" must end with</td>'
            message += '<td><input id="input" type="text" name="inputEnds"></td></tr>'
        elif thisConstraint == 'contains':
            message += '<tr><td style="font-size:150%">Enter the character must be contained in  data from column "{thisColumnName}"</td>'
            message += '<td><input id="input" type="text" name="inputContains"></td></tr>'
        elif thisConstraint == 'notContains':
            message += f'<tr><td style="font-size:150%">Enter the characters must <b>not</b> be contained in data from column "{thisColumnName}"</td>'
            message += '<td><input id="input" type="text" name="inputNotContains"></td></tr>'
        elif thisConstraint == 'inRange':
            message += f'<tr><td style="font-size:150%">Enter the minimum value for data from column "{thisColumnName}"</td>'
            message += '<td><input id="input" type="text" name="inputInRangeLow"></td></tr>'
            message += '<td><input id="lowRangeExclude" type="checkbox" name="lowRangeExclude"></td><td>Exclude this value from the mined data</td></tr>'
            message += f'<tr><td style="font-size:150%">Enter the maxumum value for data from column "{thisColumnName}"</td>'
            message += '<td><input id="input" type="text" name="inputInRangeHigh"></td></tr>'
            message += '<td><input id="highRangeExclude" type="checkbox" name="highRangeExclude"></td><td>Exclude this value from the mined data</td></tr>'
        else:
            message = f'<p style="text-align:centre"><b><a href="{url_for("splash")}">Internal error (unknown constraint type "{thisConstraint}") - please click here to start again</a></b>'
            return message
    message += '</table>'
    message += '<br/>'
    message += '<input id="submit" name="submit" type="submit" value="Set this/these constraint(s)" style="font-size:150%">'
    message += '</form>'
    message += '</body></html>'
    return message


def testValue(value, datatype):
    '''
    Test if a value matches the datatype and return an SQL valid version of value
    '''
    if datatype == 'string':
        return str(value)
    elif datatype in ['int', 'float', 'numeric', 'decimal']:
        try:
            x = float(value)
        except Exception as thisE:
            return None
        return x
    else:
        try:
            thisDatetime = dateutil.parser.parse(str(value))
        except Exception as thisE:
            return None
        if datatype == 'date':
            return thisDatetime.date().isoformat()
        else:
            return thisDatetime.isoformat()

def setValue(where, thisColumn, relop, value, datatype):
    '''
    Construct a constraint for this column
    '''
    if where is None:
        where = ''
    if where != '':
        where += ' AND '
    where += f'{thisColumn} {relop} '
    if datatype in ['int', 'float', 'numeric', 'decimal']:
        where += f'{value}'
    else:
        where += f'"{value}"'
    return where


@app.route('/setConstraints', methods=['POST'])
def setConstraints():
    '''
    Add to 'where' using the user entered constraint values
    '''
    message = '<html><head><title>Simple Data Miner</title><link rel="icon" href="data:,"></head><body style="font-size:120%">'
    message += '<h1 style="text-align:center">Simple Data Miner</h1>'
    thisMessage, thisTable, columnsSelected, constrainedColumns, nextConstraint, where = checkForm(request, 2)
    if thisMessage is not None:
        message += thisMessage
        return Response(response=message, status=400)
    whereWas = where
    thisCol = d.mineTables[thisTable]['columns'][constrainedColumns[nextConstraint]]
    thisColumn = thisCol['column']
    thisColumnName = thisCol['columnName']
    thisDatatype = thisCol['datatype']
    if 'inputEquals' in request.form:
        thisValue = convertInWeb(request.form['inputEquals'].strip())
        if (value := testValue(thisValue, thisDatatype)) is None:
            message += redoThisConstraint(thisTable, columnsSelected, constrainedColumns, nextConstraint, where, thisValue, thisDatatype, thisColumnName)
            return Response(response=message, status=200)
        where = setValue(where, thisColumn, '=', value, thisDatatype)
    if 'inputNotEquals' in request.form:
        thisValue = convertInWeb(request.form['inputNotEquals'].strip())
        if (value := testValue(thisValue, thisDatatype)) is None:
            message += redoThisConstraint(thisTable, columnsSelected, constrainedColumns, nextConstraint, where, thisValue, thisDatatype, thisColumnName)
            return Response(response=message, status=200)
        where = setValue(where, thisColumn, '!=', value, thisDatatype)
    if 'inputGtThan' in request.form:
        thisValue = convertInWeb(request.form['inputGtThan'].strip())
        if (value := testValue(thisValue, thisDatatype)) is None:
            message += redoThisConstraint(thisTable, columnsSelected, constrainedColumns, nextConstraint, where, thisValue, thisDatatype, thisColumnName)
            return Response(response=message, status=200)
        where = setValue(where, thisColumn, '>', value, thisDatatype)
    if 'inputGteThan' in request.form:
        thisValue = convertInWeb(request.form['inputGteThan'].strip())
        if (value := testValue(thisValue, thisDatatype)) is None:
            message += redoThisConstraint(thisTable, columnsSelected, constrainedColumns, nextConstraint, where, thisValue, thisDatatype, thisColumnName)
            return Response(response=message, status=200)
        where = setValue(where, thisColumn, '>=', value, thisDatatype)
    if 'inputLtThan' in request.form:
        thisValue = convertInWeb(request.form['inputLtThan'].strip())
        if (value := testValue(thisValue, thisDatatype)) is None:
            message += redoThisConstraint(thisTable, columnsSelected, constrainedColumns, nextConstraint, where, thisValue, thisDatatype, thisColumnName)
            return Response(response=message, status=200)
        where = setValue(where, thisColumn, '<', value, thisDatatype)
    if 'inputLteThan' in request.form:
        thisValue = convertInWeb(request.form['inputLteThan'].strip())
        if (value := testValue(thisValue, thisDatatype)) is None:
            message += redoThisConstraint(thisTable, columnsSelected, constrainedColumns, nextConstraint, where, thisValue, thisDatatype, thisColumnName)
            return Response(response=message, status=200)
        where = setValue(where, thisColumn, '<=', value, thisDatatype)
    if 'inputStarts' in request.form:
        thisValue = convertInWeb(request.form['inputStarts'].strip())
        if (value := testValue(thisValue, thisDatatype)) is None:
            message += redoThisConstraint(thisTable, columnsSelected, constrainedColumns, nextConstraint, where, thisValue, thisDatatype, thisColumnName)
            return Response(response=message, status=200)
        where = setValue(where, thisColumn, 'like', f'{value}%', thisDatatype)
    if 'inputEnds' in request.form:
        thisValue = convertInWeb(request.form['inputEnds'].strip())
        if (value := testValue(thisValue, thisDatatype)) is None:
            message += redoThisConstraint(thisTable, columnsSelected, constrainedColumns, nextConstraint, where, thisValue, thisDatatype, thisColumnName)
            return Response(response=message, status=200)
        where = setValue(where, thisColumn, 'like', f'%{value}', thisDatatype)
    if 'inputContains' in request.form:
        thisValue = convertInWeb(request.form['inputContains'].strip())
        if (value := testValue(thisValue, thisDatatype)) is None:
            message += redoThisConstraint(thisTable, columnsSelected, constrainedColumns, nextConstraint, where, thisValue, thisDatatype, thisColumnName)
            return Response(response=message, status=200)
        where = setValue(where, thisColumn, 'like', f'%{value}%', thisDatatype)
    if 'inputNotContains' in request.form:
        thisValue = convertInWeb(request.form['inputContains'].strip())
        if (value := testValue(thisValue, thisDatatype)) is None:
            message += redoThisConstraint(thisTable, columnsSelected, constrainedColumns, nextConstraint, where, thisValue, thisDatatype, thisColumnName)
            return Response(response=message, status=200)
        where = setValue(where, thisColumn, 'not like', f'%{value}%', thisDatatype)
    if ('inputInRangeLow' in request.form) or ('inputInRangeHigh' in request.form):
        if ('InputInRangeLow' not in request.form) or ('inputInRangeHigh' not in request.form):
            message += f'<form id="tables" action ="{url_for("doNextConstraint")}" method="post" enctype="multipart/form-data">'
            message += f'<input id="table" type="hidden" name="table" value="{thisTable}">'
            message += f'<input id="columnsSelected" type="hidden" name="columnsSelected" value="{columnsSelected}">'
            message += f'<input id="constrainedColumn" type="hidden" name="constrainedColumns" value="{constrainedColumns}">'
            message += f'<input id="nextConstraint" type="hidden" name="nextConstraint" value="{nextConstraint}">'
            message += f'<input id="where" type="hidden" name="where" value="{whereWas}">'
            message += '<input id="submit" type="submit" name="select" value="Incomplete range specification - click here to try again" style="font-size:120%">'
            message += '</form>'
            message += '</body></html>'
            return Response(response=message, status=200)
        thisValue = convertInWeb(request.form['inputInRangeLow'].strip())
        if (value := testValue(thisValue, thisDatatype)) is None:
            message += redoThisConstraint(thisTable, columnsSelected, constrainedColumns, nextConstraint, where, thisValue, thisDatatype, thisColumnName)
            return Response(response=message, status=200)
        if 'lowRangeExclude' in request.form:
            where = setValue(where, thisColumn, '>', value, thisDatatype)
        else:
            where = setValue(where, thisColumn, '>=', value, thisDatatype)
        thisValue = convertInWeb(request.form['inputInRangeHigh'].strip())
        if (value := testValue(thisValue, thisDatatype)) is None:
            message += redoThisConstraint(thisTable, columnsSelected, constrainedColumns, nextConstraint, where, thisValue, thisDatatype, thisColumnName)
            return Response(response=message, status=200)
        if 'highRangeExclude' in request.form:
            where = setValue(where, thisColumn, '<', value, thisDatatype)
        else:
            where = setValue(where, thisColumn, '<=', value, thisDatatype)
    nextConstraint += 1
    if nextConstraint < len(constrainedColumns):
        message += makeConstraint(thisTable, columnsSelected, constrainedColumns, nextConstraint, where)
    else:
        message += buildAggs(thisTable, columnsSelected, where)
    return Response(response=message, status=200)


def redoThisConstraint(thisTable, columnsSelected, constrainedColumns, nextConstraint, where, value, thisDatatype, thisColumnName):
    '''
    Redo 'doThisConstraint as the value wasn't valid
    '''
    message = f'<form id="tables" action ="{url_for("doNextConstraint")}" method="post" enctype="multipart/form-data">'
    message += f'<input id="table" type="hidden" name="table" value="{thisTable}">'
    message += f'<input id="columnsSelected" type="hidden" name="columnsSelected" value="{columnsSelected}">'
    message += f'<input id="constrainedColumn" type="hidden" name="table" value="{constrainedColumns}">'
    message += f'<input id="nextConstraint" type="hidden" name="table" value="{nextConstraint}">'
    message += f"<input id='where' type='hidden' name='where' value='{where}'>"
    message += f'<input id="submit" type="select" name="select" value="Value {value} is not valid for the datatype({thisDatatype}) for column {thisColumnName} - click here to try again">'
    message += '</form>'
    message += '</body></html>'
    return message

def buildAggs(thisTable, columnsSelected, where):
    '''
    Build the "select columns to aggregate" web page
    '''
    message = '<h2 style="text-align:center">Please select any columns you want counted and/or summed in your mined data</h2>'
    message += f'<form id="aggregates" action ="{url_for("doAggregates")}" method="post" enctype="multipart/form-data">'
    message += f'<input id="table" type="hidden" name="table" value="{thisTable}">'
    message += f'<input id="columnsSelected" type="hidden" name="columnsSelected" value="{columnsSelected}">'
    message += f"<input id='where' type='hidden' name='where' value='{where}'>"
    message += '<table>'
    message += '<tr><th style="font-size:150%">Column</th><th style="font-size:150%">count()</th><th style="font-size:150%">sum()</th></tr>'
    for thisCol in columnsSelected:
        if d.mineTables[thisTable]['columns'][thisCol]['datatype'] not in ['int', 'float', 'numeric', 'decimal']:
            continue
        columnName = d.mineTables[thisTable]["columns"][thisCol]['columnName']
        message += '<tr>'
        message += f'<td style="font-size:150%">{columnName}</td>'
        message += f'<td><input id="checked" type="checkbox" name="selectCount" value="{thisCol}"></td>'
        message += f'<td><input id="checked" type="checkbox" name="selectSum" value="{thisCol}"></td>'
        message += '</tr>'
    message += '</table>'
    message += f'<input id="submit" type="submit" name="submit" value="Please count/sum these columns in the {d.mineTables[thisTable]["tableName"]} table" style="font-size:150%">'
    message += '</form>'
    message += '</body></html>'
    return message

@app.route('/doAggregates', methods=['POST'])
def doAggregates():
    '''
    Implement any summing, counting and grouping
    '''
    message = '<html><head><title>Simple Data Miner</title><link rel="icon" href="data:,"></head><body style="font-size:120%">'
    message += '<h1 style="text-align:center">Simple Data Miner</h1>'
    thisMessage, thisTable, columnsSelected, dummy1, dummy2, where = checkForm(request, 3)
    if thisMessage is not None:
        message += thisMessage
        return Response(response=message, status=400)
    where = convertInWeb(request.form['where'].strip())
    selectColumns = ''
    groupByColumns = ''
    countThese = []
    sumThese = []
    if ('selectCount' in request.form) or ('selectSum' in request.form):
        if 'selectCount' in request.form:
            for countIt in request.form.getlist('selectCount'):
                countThese.append(int(countIt))
        if 'selectSum' in request.form:
            for sumIt in request.form.getlist('selectSum'):
                sumThese.append(int(sumIt))
        for col in columnsSelected:
            thisCol = d.mineTables[thisTable]['columns'][int(col)]
            thisColumn = thisCol['column']
            if (col in countThese) or (col in sumThese):
                if col in countThese:
                    if selectColumns != '':
                        selectColumns += ', '
                    selectColumns += f'count({thisColumn})'
                if col in sumThese:
                    if selectColumns != '':
                        selectColumns += ', '
                    selectColumns += f'sum({thisColumn})'
            else:
                if selectColumns != '':
                    selectColumns += ', '
                selectColumns += f'{thisColumn}'
                if groupByColumns != '':
                    groupByColumns += ', '
                groupByColumns += f'{thisColumn}'
    else:
        for col in columnsSelected:
            thisCol = d.mineTables[thisTable]['columns'][int(col)]
            thisColumn = thisCol['column']
            if selectColumns != '':
                selectColumns += ', '
            selectColumns += f'{thisColumn}'
    message = '<h2 style="text-align:center">Here is your SQL query for mining your extract</h2>'
    selectText = f'SELECT {selectColumns}\nFROM {thisTable}'
    if (where is not None) and (where != ''):
        thisWhere = where.replace(' AND ','\n      AND ')
        selectText += f'\nWHERE {thisWhere}'
    if groupByColumns != '':
        selectText += f'\nGROUP BY {groupByColumns}'
    message += f'<br/><pre style="font-size:150%">{selectText}</pre>'
    message += '<br/>'
    countSelectText = f'SELECT count(*) as count FROM {thisTable}'
    if (where is not None) and (where != ''):
        countSelectText += f' WHERE {where}'
    extractCount_df = pd.read_sql_query(text(countSelectText), d.engine.connect())
    for rowCount in extractCount_df.itertuples():
        if rowCount.count > d.mineTables[thisTable]['maxRecords']:
            message += f'<p style="text-align:centre"><b><a href="{url_for("splash")}">Your mined extract would access too many records "{rowCount.count}" [limit:{d.mineTables[thisTable]["maxRecords"]}] - please click here to start again</a></b>'
            message += '</body></html>'
            return Response(response=message, status=400)
    selectText = f'SELECT {selectColumns} FROM {thisTable}'
    if (where is not None) and (where != ''):
        selectText += f' WHERE {where}'
    if groupByColumns != '':
        selectText += f' GROUP BY {groupByColumns}'
    message += f'<p style="font-size:150%"><b><a href="{url_for("doSQL", SQL=selectText)}">Click here to execute this SQL, mine your extract and download it</a></b>'
    message += f'<p style="font-size:150%"><b><a href="{url_for("splash")}">Click here to start a new data mining operation</a></b>'
    message += '</body></html>'
    return Response(response=message, status=200)


@app.route('/doSQL/<SQL>', methods=['GET'])
def doSQL(SQL):
    '''
    Execute the SQL and download the resulting Excel workbook
    '''
    extract_df = pd.read_sql_query(text(SQL), d.engine.connect())
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        extract_df.to_excel(writer, index=False)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name='SimpleDataMinerExtract.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


if __name__ == '__main__':

    '''
    The main code
    Start by parsing the command line arguements, setting up logging and connecting to the database.
    Then check the Excel workbook - it should have one sheet of database table names/worksheet names pairs.
    Then one sheet per database table of database table configuration.
    '''

    # Save the program name
    progName = sys.argv[0]
    progName = progName[0:-3]        # Strip off the .py ending

    # set the options
    parser = argparse.ArgumentParser(description='Simple Data Miner')
    parser.add_argument('-D', '--DatabaseType', dest='DatabaseType', choices=['MSSQL', 'MySQL'],
                        help='The Database Type [choices: MSSQL/MySQL]')
    parser.add_argument('-I', '--inputDir', dest='inputDir', default='.',
                        help='The directory containing the Excel workbook containing the configuration of the mineable tables.')
    parser.add_argument('-i', '--inputWorkbook', dest='inputWorkbook', default='tablesConfig.xlsx',
                        help='The name of the Excel workbook containing configuration of the mineable tables')
    parser.add_argument('-C', '--configDir', dest='configDir', default='databaseConfig',
                        help='The name of the directory containing the database connection configuration file (default=databaseConfig)')
    parser.add_argument('-c', '--configFile', dest='configFile', default='SimpleDataMiner.json',
                        help='The name of the configuration file (default SimpleDataMiner.json)')
    parser.add_argument('-s', '--server', dest='server', help='The address of the database server')
    parser.add_argument('-u', '--username', dest='username', help='The user required to access the database')
    parser.add_argument('-p', '--password', dest='password', help='The user password required to access the database')
    parser.add_argument('-d', '--databaseName', dest='databaseName', help='The name of the database')
    parser.add_argument ('-v', '--verbose', dest='verbose', type=int, choices=range(0,5), help='The level of logging\n\t0=CRITICAL,1=ERROR,2=WARNING,3=INFO,4=DEBUG')
    parser.add_argument ('-L', '--logDir', dest='logDir', default='.', metavar='logDir', help='The name of the directory where the logging file will be created')
    parser.add_argument ('-l', '--logFile', dest='logFile', metavar='logfile', help='The name of a logging file')
    args = parser.parse_args()

    # Parse the command line options
    args = parser.parse_args()
    DatabaseType = args.DatabaseType
    inputDir = args.inputDir
    inputWorkbook = args.inputWorkbook
    configDir = args.configDir
    configFile = args.configFile
    server = args.server
    username = args.username
    password = args.password
    databaseName = args.databaseName
    logDir = args.logDir
    logFile = args.logFile
    loggingLevel = args.verbose

    # Set up logging
    logging_levels = {0:logging.CRITICAL, 1:logging.ERROR, 2:logging.WARNING, 3:logging.INFO, 4:logging.DEBUG}
    logfmt = progName + ' [%(asctime)s]: %(message)s'
    if loggingLevel is not None:    # Change the logging level from "WARN" if the -v vebose option is specified
        if logFile is not None:        # and send it to a file if the -o logfile option is specified
            with open(os.path.join(logDir, logFile), 'wt', encoding='utf-8', newline='') as logOutput:
                pass
            logging.basicConfig(format=logfmt, datefmt='%d/%m/%y %H:%M:%S %p', level=logging_levels[loggingLevel], filename=os.path.join(logDir, logFile))
        else:
            logging.basicConfig(format=logfmt, datefmt='%d/%m/%y %H:%M:%S %p', level=logging_levels[loggingLevel])
    else:
        if logFile is not None:        # send the default (WARN) logging to a file if the -o logfile option is specified
            with open(os.path.join(logDir, logFile), 'wt', encoding='utf-8', newline='') as logOutput:
                pass
            logging.basicConfig(format=logfmt, datefmt='%d/%m/%y %H:%M:%S %p', filename=os.path.join(logDir, logFile))
        else:
            logging.basicConfig(format=logfmt, datefmt='%d/%m/%y %H:%M:%S %p')

    # Set the SQLAlchemy logging level
    if loggingLevel is not None:
        logging.getLogger('sqlalchemy.engine').setLevel(logging_levels[loggingLevel])
    else:
        logging.getLogger('sqlalchemy.engine').setLevel(logging.WARN)

    # Read in the configuration file - which must exist if required - and create the database engine
    config = {}                 # The configuration data
    try:
        with open(os.path.join(configDir, configFile), 'rt', newline='', encoding='utf-8') as configSource:
            config = json.load(configSource, object_pairs_hook=collections.OrderedDict)
    except IOError:
        logging.critical('configFile (%s/%s) failed to load', configDir, configFile)
        logging.shutdown()
        sys.exit(d.EX_CONFIG)

    # Check that we have a databaseName if we have a databaseType
    if DatabaseType not in config:
        logging.critical('DatabaseType(%s) not found in configuraton file(%s)', DatabaseType, configFile)
        logging.shutdown()
        sys.exit(d.EX_USAGE)
    if 'connectionString' not in config[DatabaseType]:
        logging.critical('No %s connectionString defined in configuration file(SQLAlchemyDB.json)', DatabaseType)
        logging.shutdown()
        sys.exit(d.EX_CONFIG)
    connectionString = config[DatabaseType]['connectionString']
    if ('username' in config[DatabaseType]) and (username is None):
        username = config[DatabaseType]['username']
    if ('password' in config[DatabaseType]) and (password is None):
        password = config[DatabaseType]['password']
    if ('server' in config[DatabaseType]) and (server is None):
        server = config[DatabaseType]['server']
    if ('databaseName' in config[DatabaseType]) and (databaseName is None):
        databaseName = config[DatabaseType]['databaseName']

    # Check that we have all the required paramaters
    if username is None:
        logging.critical('Missing definition for "username"')
        logging.shutdown()
        sys.exit(d.EX_USAGE)
    if password is None:
        logging.critical('Missing definition for "password"')
        logging.shutdown()
        sys.exit(d.EX_USAGE)
    if server is None:
        logging.critical('Missing definition for "server"')
        logging.shutdown()
        sys.exit(d.EX_USAGE)
    if databaseName is None:
        logging.critical('Missing definition for "databaseName"')
        logging.shutdown()
        sys.exit(d.EX_USAGE)
    connectionString = connectionString.format(username=username, password=password, server=server, databaseName=databaseName)

    # Create the engine
    if DatabaseType == 'MSSQL':
        d.engine = create_engine(connectionString, use_setinputsizes=False, echo=False)
    else:
        d.engine = create_engine(connectionString, echo=False)

    # Check if the database exists
    if not database_exists(d.engine.url):
        logging.critical('Database %s does not exist', databaseName)
        logging.shutdown()
        sys.exit(d.EX_CONFIG)

    # Connect to the database
    try:
        conn = d.engine.connect()
    except OperationalError:
        logging.critical('Connection error for database %s', databaseName)
        logging.shutdown()
        sys.exit(d.EX_UNAVAILABLE)
    except Exception as e:
        logging.critical('Connection error for database %s:%s', databaseName, e.args)
        logging.shutdown()
        sys.exit(d.EX_UNAVAILABLE)
    conn.close()

    # Now get the metadata and build a session maker
    d.metadata = MetaData()
    d.metadata.reflect(bind=d.engine, views=True)
    d.Session = sessionmaker(bind=d.engine)


    # Load the configuration workbook
    wb = load_workbook(os.path.join(inputDir, inputWorkbook))

    # Check the 'tables' worksheet
    if 'tables' not in wb.sheetnames:
        logging.critical('No sheet name "tables" in workbook')
        logging.shutdown()
        sys.exit(d.EX_CONFIG)
    ws = wb['tables']
    data = ws.values
    cols = next(data)
    tables_df = pd.DataFrame(list(data), columns=cols)
    if 'table' not in tables_df.columns:
        logging.critical('Missing "table" heading in "tables" worksheet')
        logging.shutdown()
        sys.exit(d.EX_CONFIG)
    if 'tableName' not in tables_df.columns:
        logging.critical('Missing "tableName" heading in "tables" worksheet')
        logging.shutdown()
        sys.exit(d.EX_CONFIG)
    if 'worksheet' not in tables_df.columns:
        logging.critical('Missing "worksheet" heading in "tables" worksheet')
        logging.shutdown()
        sys.exit(d.EX_CONFIG)
    if 'maxRecords' not in tables_df.columns:
        logging.critical('Missing "maxRecords" heading in "tables" worksheet')
        logging.shutdown()
        sys.exit(d.EX_CONFIG)
    tables = {}
    for tableRow in tables_df.itertuples():
        table = tableRow.table
        tableName = tableRow.tableName
        worksheet = tableRow.worksheet
        maxRecords = tableRow.maxRecords
        # Check this table exists
        if table not in d.metadata.tables:
            logging.critical('Table "%s" not in database', table)
            logging.shutdown()
            sys.exit(d.EX_CONFIG)
        # Check that this worksheet exits
        if worksheet not in wb.sheetnames:
            logging.critical('No sheet named "%s" in workbook', worksheet)
            logging.shutdown()
            sys.exit(d.EX_CONFIG)

        # Check this worksheet
        ws = wb[worksheet]
        data = ws.values
        cols = next(data)
        thisTable_df = pd.DataFrame(list(data), columns=cols)
        if 'column' not in thisTable_df.columns:
            logging.critical('Missing "column" heading in "%s" worksheet', worksheet)
            logging.shutdown()
            sys.exit(d.EX_CONFIG)
        if 'columnName' not in thisTable_df.columns:
            logging.critical('Missing "columnName" heading in "%s" worksheet', worksheet)
            logging.shutdown()
            sys.exit(d.EX_CONFIG)
        if 'datatype' not in thisTable_df.columns:
            logging.critical('Missing "datatype" heading in "%s" worksheet', worksheet)
            logging.shutdown()
            sys.exit(d.EX_CONFIG)
        if 'isIndexed' not in thisTable_df.columns:
            logging.critical('Missing "isIndexed" heading in "%s" worksheet', worksheet)
            logging.shutdown()
            sys.exit(d.EX_CONFIG)
        if 'lookupTable' not in thisTable_df.columns:
            logging.critical('Missing "lookupTable" heading in "%s" worksheet', worksheet)
            logging.shutdown()
            sys.exit(d.EX_CONFIG)
        if 'lookupCodeColumn' not in thisTable_df.columns:
            logging.critical('Missing "lookupCodeColumn" heading in "%s" worksheet', worksheet)
            logging.shutdown()
            sys.exit(d.EX_CONFIG)
        if 'lookupDescriptionColumn' not in thisTable_df.columns:
            logging.critical('Missing "lookupDescriptionColumn" heading in "%s" worksheet', worksheet)
            logging.shutdown()
            sys.exit(d.EX_CONFIG)

        d.mineTables[table] = {}
        d.mineTables[table]['tableName'] = tableName
        d.mineTables[table]['maxRecords'] = maxRecords
        d.mineTables[table]['columns'] = []
        for columnRow in thisTable_df.itertuples():
            if columnRow.column not in d.metadata.tables[table].columns:
                logging.critical('No column named "%s" in table "%s" not in database', columnRow.column, table)
                logging.shutdown()
                sys.exit(d.EX_CONFIG)
            if columnRow.lookupTable is not None:
                if columnRow.lookupTable not in d.metadata.tables:
                    logging.critical('Table "%s" not in database', columnRow.lookupTable)
                    logging.shutdown()
                    sys.exit(d.EX_CONFIG)
                if columnRow.lookupCodeColumn not in d.metadata.tables[columnRow.lookupTable].columns:
                    logging.critical('No column named "%s" in table "%s" not in database', columnRow.lookupCodeColumn, columnRow.lookupTable)
                    logging.shutdown()
                    sys.exit(d.EX_CONFIG)
                if columnRow.lookupDescriptionColumn not in d.metadata.tables[columnRow.lookupTable].columns:
                    logging.critical('No column named "%s" in table "%s"', columnRow.lookupDescriptionColumn, columnRow.lookupTable)
                    logging.shutdown()
                    sys.exit(d.EX_CONFIG)
            column = {}
            column['column'] = columnRow.column
            column['columnName'] = columnRow.columnName
            column['datatype'] = columnRow.datatype
            if column['datatype'] not in ['string', 'int', 'float', 'numeric', 'decimal', 'date', 'datetime']:
                logging.critical('Invalid datatype "%s" for column "%s" in table "%s"', columnRow.datatype, columnRow.column, table)
                logging.critical('Must be one of "string", "int", "float", "numeric", "decimal", "date", "datetime", "time"')
                logging.shutdown()
                sys.exit(d.EX_CONFIG)
            column['isIndexed'] = columnRow.isIndexed
            column['lookupTable'] = columnRow.lookupTable
            column['lookupCodeColumn'] = columnRow.lookupCodeColumn
            column['lookupDescriptionColumn'] = columnRow.lookupDescriptionColumn
            d.mineTables[table]['columns'].append(column)

    app.run(host="0.0.0.0")
